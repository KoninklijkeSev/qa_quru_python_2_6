import os.path
import csv
import zipfile
from zipfile import ZipFile
from PyPDF2 import PdfReader
from openpyxl import load_workbook

def test_archive():
    docs_pdf = os.path.abspath('./resources/pdf.pdf')
    docs_csv = os.path.abspath('./resources/csv.csv')
    docs_xlsx = os.path.abspath('./resources/xlsx.xlsx')
    print(docs_pdf)
    print(docs_csv)
    print(docs_xlsx)

    zip_archive = ZipFile('Archive.zip', 'w')
    zip_archive.write(docs_xlsx, arcname='xlsx_added.xlsx')
    zip_archive.write(docs_csv, arcname='csv_added.csv')
    zip_archive.write(docs_pdf, arcname='pdf_added.pdf')

    zip_archive.close()

def test_check_pdf():
    pdf_reader = PdfReader('./resources/pdf.pdf')
    number_of_pages = len(pdf_reader.pages)
    assert number_of_pages == 2
    page = pdf_reader.pages[0]
    text = page.extract_text()
    assert 'Python 3 Cheat Sheet' in text


def test_check_csv():
    with open('./resources/csv.csv') as csvfile:
        table = csv.reader(csvfile)
        for line_no, column in enumerate(table, 0):
            if line_no == 1:
                assert '2016-01-01' in column[1]

def test_check_xlsx():
    workbook = load_workbook('./resources/xlsx.xlsx')
    sheet = workbook.active
    assert 'France' == sheet.cell(row=4, column=5).value

